"""
InlineComp – Deelnemerslijst-diff (v5+, gebouwd in stijl van v4)

Vergelijkt twee deelnemerslijsten voor een wedstrijd:
  1) InlineComp-PDF (Deelnemerslijst-print uit Print Center)
  2) Organisatie-Excel (aanmeldingen-overzicht)

Output: opgemaakte PDF met:
  - Samenvattings-tabel (categorieën + counts)
  - 1. Alleen in Excel (vermoedelijk vergeten in InlineComp)
  - 2. Alleen in InlineComp (vermoedelijk afgemeld of niet bij org bekend)
  - 3. Naam-verschillen (fuzzy-match met match-percentage)
  - 4. Afstand-verschillen per rijder (PDF=ja/nee, XLS=ja/nee)
  - 5. Startnummer-verschillen op gemeenschappelijke namen
  - Actie-checklist voor de organisatie
  - Bronbestanden-referentie

Match-flow:
  1. Exact op (snr + genormaliseerde naam)
  2. Fuzzy op naam (difflib.SequenceMatcher, ≥85% voor "vermoedelijk zelfde")
  3. Snr-mismatch op exact-naam-match (zelfde naam, ander snr)

Usage:
    python tools/deelnemers_diff.py <inlinecomp.pdf> <organisatie.xlsx> [output.pdf] [--sheet=<naam>]

  --sheet=<naam>  Forceer Excel-sheet-naam (anders interactieve keuze bij
                  meerdere data-sheets).
"""
from __future__ import annotations
import re
import sys
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook
import pdfplumber
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Table, TableStyle,
                                Spacer, KeepTogether)
from reportlab.lib import colors


# ── Naam-normalisatie ────────────────────────────────────────────────────
def normaliseer_naam(s: str) -> str:
    s = (s or '').lower().strip()
    s = re.sub(r"[''`]", '', s)
    s = re.sub(r'\s+', ' ', s)
    return s


def fuzzy_score(a: str, b: str) -> int:
    na, nb = normaliseer_naam(a), normaliseer_naam(b)
    # Word-set match: "elbrich nicolay" en "nicolay elbrich" = 100%
    set_score = round(
        SequenceMatcher(None,
                        ' '.join(sorted(na.split())),
                        ' '.join(sorted(nb.split()))).ratio() * 100)
    seq_score = round(SequenceMatcher(None, na, nb).ratio() * 100)
    return max(set_score, seq_score)


# ── Afstand-normalisatie ────────────────────────────────────────────────
def normaliseer_afstand(token: str) -> str:
    """
    Reduceert PDF/Excel-tekst naar één canoniek label per afstand.
    Voorbeelden:
      "Cadet Ladies 200m DTT"   → "200m"
      "1000m"                    → "1000m"
      "Pointsrace"               → "puntenkoers"
      "Puntenkoers 3000m"        → "puntenkoers"
      "Afvalkoers 5000m"         → "afvalkoers"
      "Aflossing 3000m"          → "aflossing"
    """
    t = (token or '').lower().strip()
    if not t:
        return ''
    if re.search(r'\baflossing\b|relay', t):                  return 'aflossing'
    if re.search(r'\bafvalkoers\b|elimination|pursuit', t):   return 'afvalkoers'
    if re.search(r'\bpuntenkoers\b|points\s?race|pointsrace', t): return 'puntenkoers'
    if re.search(r'\b200\s?m?\b|dtt|dual', t):                return '200m'
    if re.search(r'\b500\s?m?\b', t):                         return '500m'
    if re.search(r'\b1000\s?m?\b', t):                        return '1000m'
    if re.search(r'\b3000\s?m\b', t):                         return '3000m'
    if re.search(r'\b5000\s?m\b', t):                         return '5000m'
    if re.search(r'\b10000\s?m\b', t):                        return '10000m'
    return t


# Mapping van Excel-header-aliassen naar canonieke afstand-namen.
# OH850 (en mogelijk andere wedstrijden van Skeelerclub Heerde): de organisatie-
# Excel gebruikt de kolom 'afvalkoers' historisch voor wat in InlineComp
# puntenkoers heet. Geert heeft dit meerdere keren bevestigd — sla het hier
# expliciet op zodat het niet verkeerd geïnterpreteerd wordt.
EXCEL_HEADER_ALIAS = {
    'afvalkoers': 'puntenkoers',
}


# ── Excel parser (v4-stijl, exacte kolom-indexes) ───────────────────────
def _kies_excel_sheet(wb, voorkeur: str | None = None) -> str:
    """
    Kies welke sheet uit de werkboek-file gebruikt wordt:
      - Als `voorkeur` (uit --sheet argument) gegeven en bestaat → die.
      - Anders: alle sheets met meer dan 30 niet-lege rijen tellen mee.
        - Eén kandidaat → die.
        - Meerdere → vraag interactief.
        - Geen → eerste sheet uit de file als fallback.
    """
    if voorkeur:
        v_lower = voorkeur.lower()
        for sname in wb.sheetnames:
            if sname.lower() == v_lower:
                return sname

    kandidaten = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        non_empty = 0
        for r in ws.iter_rows(values_only=True):
            if r and any(c is not None and str(c).strip() for c in r):
                non_empty += 1
            if non_empty > 30:
                break
        if non_empty > 30:
            kandidaten.append((sname, non_empty))
    if not kandidaten:
        return wb.sheetnames[0]
    if len(kandidaten) == 1:
        return kandidaten[0][0]

    # Meerdere data-sheets: interactief kiezen
    print('Excel bevat meerdere data-sheets — kies welke:')
    for i, (s, n) in enumerate(kandidaten, 1):
        ws = wb[s]
        # Tel alle rijen voor display
        total = sum(1 for r in ws.iter_rows(values_only=True)
                    if r and any(c is not None and str(c).strip() for c in r))
        print(f'  [{i}] {s} ({total} rijen)')
    while True:
        try:
            keuze = input(f'Sheet-nummer (1-{len(kandidaten)}) of naam: ').strip()
        except EOFError:
            return kandidaten[0][0]
        if keuze.isdigit() and 1 <= int(keuze) <= len(kandidaten):
            return kandidaten[int(keuze) - 1][0]
        if keuze in wb.sheetnames:
            return keuze
        print('Ongeldige keuze, probeer opnieuw.')


def parse_excel(path: Path, sheet: str | None = None) -> tuple[list[dict], list[str]]:
    """
    Leest de organisatie-Excel zoals OH850's "aanmeldingen bijgewerkt"-sheet:
    16 kolommen, vaste indices:
      0 vn | 1 tu | 2 an | 3 volledige_naam | 4 land_lang | 5 land |
      6 geslacht (M/W) | 7 cat (Pupil/Cadet/Youth/Junior/Senior) |
      8 _ | 9 startnr | 10 _ | 11 club | 12 200m | 13 1000m |
      14 'afvalkoers' (= eigenlijk puntenkoers in deze Excel) | 15 _

    EXCEL_HEADER_ALIAS zorgt dat kolom 14 ('afvalkoers') naar 'puntenkoers' mapped.

    Sheet-keuze:
      - `sheet`-parameter (uit --sheet argument) heeft voorrang.
      - Anders: als de werkboek meerdere data-sheets heeft, vraagt
        het script interactief welke gebruikt moet worden.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_naam = _kies_excel_sheet(wb, voorkeur=sheet)
    print(f'Excel-sheet: {sheet_naam}')
    ws = wb[sheet_naam]

    deelnemers = []
    afstand_keys = ['200m', '1000m', 'puntenkoers']
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or len(r) < 15:
            continue
        vn       = r[0] or ''
        tu       = r[1] or ''
        an       = r[2] or ''
        comp     = r[3]
        gesl     = (str(r[6]).strip() if r[6] else '')
        cat_raw  = (str(r[7]).strip() if r[7] else '')
        snr_raw  = r[9]
        club_raw = r[11]
        d200_v   = r[12]
        d1000_v  = r[13]
        dpunt_v  = r[14]

        # Naam: voorkeur voor 'volledige_naam'-kolom, anders samenstellen
        if comp and str(comp).strip():
            naam = str(comp).strip()
        else:
            naam = ' '.join(str(x).strip() for x in (vn, tu, an) if x).strip()
        if not naam:
            continue
        # Geslacht-kolom is in deze Excel soms gevuld met alleen spaties —
        # rij is dan toch een valide deelnemer. Niet skippen.
        gesl = gesl.strip()[:1].upper() if gesl else ''
        # Filter losse commentaar-regels (bv. "Rood: betaling is onderweg"
        # ergens in de naam-kolom). Geldige deelnemers hebben een bekende
        # cat-categorie (Pupil/Cadet/Youth/Junior/Senior).
        VALID_EXCEL_CATS = {k[0] for k in EXCEL_CAT_MAP.keys()}
        if cat_raw not in VALID_EXCEL_CATS:
            continue
        # Cat-mapping faalt zonder geslacht: gebruik raw cat als fallback
        cat = EXCEL_CAT_MAP.get((cat_raw, gesl)) or cat_raw or '?'

        snr = int(snr_raw) if isinstance(snr_raw, (int, float)) and snr_raw else None
        club = str(club_raw).strip() if club_raw else ''

        # Kolom P (index 15) bevat 'transponder' voor rijders ZONDER
        # geregistreerde transponder (organisatie-notatie).
        kolom_p = r[15] if len(r) > 15 else None
        geen_transponder = (isinstance(kolom_p, str)
                            and kolom_p.strip().lower() == 'transponder')

        afstanden = set()
        if isinstance(d200_v, str) and d200_v.strip().lower() in ('x', 'ja'):
            afstanden.add('200m')
        elif d200_v in (1, True):
            afstanden.add('200m')
        if isinstance(d1000_v, str) and d1000_v.strip().lower() in ('x', 'ja'):
            afstanden.add('1000m')
        elif d1000_v in (1, True):
            afstanden.add('1000m')
        # Kolom 14 heet 'afvalkoers' in de Excel maar IS puntenkoers
        if isinstance(dpunt_v, str) and dpunt_v.strip().lower() in ('x', 'ja'):
            afstanden.add('puntenkoers')
        elif dpunt_v in (1, True):
            afstanden.add('puntenkoers')

        deelnemers.append({
            'snr':   snr,
            'naam':  naam,
            'cat':   cat,
            'club':  club,
            'afstanden': afstanden,
            'geslacht_ontbreekt': (not gesl),
            'geen_transponder': geen_transponder,
        })
    return deelnemers, afstand_keys


# ── PDF parser ───────────────────────────────────────────────────────────
CAT_RX = re.compile(r'\b([DH](?:P[1-9]|K[A-Z]|J[A-Z]|S[A-Z]?|\d{2}))\b')


# Cat-mapping voor OH850-Excel: (raw_cat, geslacht) → InlineComp 3-letter code.
# Excel gebruikt 'Pupil'/'Cadet'/'Youth'/'Junior'/'Senior' + 'M'/'W'.
EXCEL_CAT_MAP = {
    ('Pupil',  'M'): 'HP1',  ('Pupil',  'W'): 'DP1',
    ('Cadet',  'M'): 'HKA',  ('Cadet',  'W'): 'DKA',
    ('Youth',  'M'): 'HJB',  ('Youth',  'W'): 'DJB',
    ('Junior', 'M'): 'HJA',  ('Junior', 'W'): 'DJA',
    ('Senior', 'M'): 'HSA',  ('Senior', 'W'): 'DSA',
}


def _rijen_op_y(words, y_tolerantie: float = 3.0):
    """Groepeer pdfplumber-words op vergelijkbare top-y (= zelfde tabelrij)."""
    if not words:
        return []
    gesort = sorted(words, key=lambda w: (round(w['top'], 1), w['x0']))
    rijen, huidig, huidig_top = [], [], None
    for w in gesort:
        if huidig_top is None or abs(w['top'] - huidig_top) <= y_tolerantie:
            huidig.append(w)
            huidig_top = w['top'] if huidig_top is None else huidig_top
        else:
            rijen.append((huidig_top, huidig))
            huidig, huidig_top = [w], w['top']
    if huidig:
        rijen.append((huidig_top, huidig))
    # Sorteer per rij links → rechts
    return [(y, sorted(rij, key=lambda w: w['x0'])) for y, rij in rijen]


def _detect_headers_per_pagina(rijen):
    """
    Zoek op een pagina de tabel-header en geef terug:
      - afstand_kolommen: lijst van (x_centrum, canoniek_label)
      - data_start_y: y onder de header van waar de rijen beginnen
    Geeft (None, None) als geen tabel-header gevonden.
    """
    for y, rij in rijen:
        teksten = [w['text'] for w in rij]
        joined = ' '.join(teksten).lower()
        # Header herkenning: bevat # + naam + cat + transp ergens
        if ('#' in teksten and 'naam' in joined and 'cat' in joined):
            # Vind de afstand-kolommen rechts van transp
            afst_cols = []
            for w in rij:
                t = w['text']
                canon = normaliseer_afstand(t)
                if canon in ('200m', '500m', '1000m', '3000m', '5000m',
                             '10000m', 'puntenkoers', 'afvalkoers', 'aflossing'):
                    afst_cols.append((w['x0'] + w['width'] / 2, canon))
            if afst_cols:
                return afst_cols, y + 5
            # Header zonder afstand-kolommen ("Transponder (invullen)") —
            # data hieronder heeft geen afstand-checkmarks
            return [], y + 5
    return None, None


def _detect_sectie_voor_y(rijen, header_y):
    """Loop boven de header en pak de eerstgevonden sectie-indicator."""
    boven_header = [(y, r) for y, r in rijen if y < header_y]
    for y, rij in reversed(boven_header):
        tekst = ' '.join(w['text'] for w in rij).lower()
        if 'afgemeld' in tekst or 'niet aanwezig' in tekst: return 'afgemeld'
        if 'door organisatie' in tekst:                     return 'org_toegevoegd'
        if 'door rijder' in tekst or 'bevestigd' in tekst:
            if 'niet bevestigd' in tekst:                   return 'niet_bevestigd'
            return 'bevestigd'
    return 'onbekend'


def _parse_data_rij(rij_words, afst_cols, tekst_sectie: bool):
    """
    Parse één tabel-rij. Retourneert (deelnemer-dict) of None.
    Bij tekst_sectie=True: zoek afstand-tokens in de naam-na-cat tekst
    (de p1 secties "Door organisatie toegevoegd" hebben afstand-namen in tekst).
    Bij tekst_sectie=False: lees ✕-checkmarks per afst-kolom.
    """
    if not rij_words:
        return None
    snr_w = rij_words[0]
    if not snr_w['text'].isdigit():
        return None
    snr = int(snr_w['text'])
    if not (1 <= snr <= 9999):
        return None
    # Zoek cat-token in rij
    cat = None
    cat_idx = None
    for i, w in enumerate(rij_words):
        m = CAT_RX.fullmatch(w['text'])
        if m:
            cat = m.group(1)
            cat_idx = i
            break
    if cat is None:
        return None
    # Naam = woorden tussen snr en cat (skip eventuele tussenliggende getallen)
    naam = ' '.join(w['text'] for w in rij_words[1:cat_idx]).strip()
    afstanden = set()

    if afst_cols:
        # Coord-based: per checkmark kijken in welke kolom 'ie valt
        for w in rij_words[cat_idx + 1:]:
            if w['text'] in ('✕', '✓', '×'):
                wc = w['x0'] + w['width'] / 2
                # Pak dichtstbijzijnde afst-kolom
                beste = min(afst_cols, key=lambda c: abs(c[0] - wc))
                if abs(beste[0] - wc) < 25:
                    afstanden.add(beste[1])

    # Detecteer "Afgem." / "afgemeld" inline → markeer als afgemeld
    rij_tekst = ' '.join(w['text'] for w in rij_words).lower()
    is_afgemeld = 'afgem' in rij_tekst or 'withdrawn' in rij_tekst

    # Tekstuele afstand-detectie (voor p1 secties zonder checkmarks)
    if tekst_sectie or not afst_cols:
        na_cat = ' '.join(w['text'] for w in rij_words[cat_idx + 1:])
        for token in re.findall(
                r'(?:200m\s*(?:dtt|dual\s*tijdrit)?|500m\+?d?|1000m|'
                r'3000m|5000m|10000m|aflossing|puntenkoers|points\s?race|'
                r'pointsrace|afvalkoers|relay|elimination|pursuit)',
                na_cat, re.IGNORECASE):
            canon = normaliseer_afstand(token)
            if canon:
                afstanden.add(canon)

    return {
        'snr': snr, 'naam': naam, 'cat': cat,
        'club': '', 'afstanden': afstanden,
        'is_afgemeld': is_afgemeld,
    }


def _extract_pdf_datum(path: Path) -> str:
    """
    Pak 'Stand: dd-mm-jjjj' (of dd-mm-jjjj, hh:mm) uit de PDF-header.
    Wordt als datum-stempel gebruikt in de rapport-koppen.
    """
    try:
        with pdfplumber.open(path) as pdf:
            t = (pdf.pages[0].extract_text() or '') if pdf.pages else ''
        m = re.search(r'Stand:\s*(\d{1,2}-\d{1,2}-\d{2,4})', t)
        if m:
            return m.group(1)
    except Exception:
        pass
    return ''


def parse_pdf(path: Path) -> list[dict]:
    """
    Parser-stijl identiek aan de v4-build:
      - extract_words(use_text_flow=False)
      - Groepeer per regel (round(top))
      - Vind tabel-header: "200m" + "1000m" + "Pointsrac" en ("Transp" of "Name")
      - Capture kolom-x0 van 200m / 1000m / Pointsrac (= 200m / 1000m / puntenkoers)
      - Daarna: alleen regels die met startnr beginnen → data-rij
      - Per rij: tellingen ✕/✗ words, bepaal dichtstbijzijnde kolom
      - Eerste hit per genormaliseerde naam wint (geen merge)

    NB: secties op p1 ("Niet aanwezig" / "Door organisatie toegevoegd")
    worden bewust NIET geparsed — de Skater-list tabel bevat sowieso alle
    rijders die in InlineComp actief zijn.
    """
    CAT_PATT = re.compile(r'^[DH](?:KA|JA|JB|SA|SB|SJ|P[1-9])$')
    pdf_d: dict[str, dict] = {}  # genormaliseerde naam → rij
    in_skater_list = False
    col_200 = col_1000 = col_punt = None
    skater_list_pages: set[int] = set()

    # Vooraf scannen: op welke pagina + y-positie staat "Deelnemerslijst
    # nnn deelnemers"? Alles VOOR die positie skippen we (= "Niet aanwezig",
    # "Door organisatie toegevoegd", "Niet bevestigd", "Race-groepen overzicht").
    deelnemerslijst_start_page: int | None = None
    deelnemerslijst_start_y: float = 0.0
    with pdfplumber.open(path) as doc:
        for pidx, page in enumerate(doc.pages):
            for w in page.extract_words(use_text_flow=False):
                if w['text'].lower().startswith('deelnemerslijst'):
                    deelnemerslijst_start_page = pidx
                    deelnemerslijst_start_y = w['top']
                    break
            if deelnemerslijst_start_page is not None:
                break

    with pdfplumber.open(path) as doc:
        for pidx, page in enumerate(doc.pages):
            # Pas beginnen bij de pagina waar "Deelnemerslijst nnn deelnemers"
            # staat. Alles ervoor (Niet aanwezig / Door organisatie / Niet
            # bevestigd / Race-groepen overzicht) is GEEN echte deelnemerslijst.
            if deelnemerslijst_start_page is not None and pidx < deelnemerslijst_start_page:
                continue
            words = page.extract_words(use_text_flow=False)
            # Op de start-pagina: filter alleen woorden ná de "Deelnemerslijst"-y
            if pidx == deelnemerslijst_start_page:
                words = [w for w in words if w['top'] >= deelnemerslijst_start_y]
            # Pre-scan voor kolom-headers: 200m / 1000m / Pointsrac kunnen op
            # verschillende y-regels staan (door "200m DTT" wrap). Pak hun
            # x0 ergens uit de pagina-header (binnen de eerste 80px na de start).
            header_y_max = None
            min_top = min((w['top'] for w in words), default=0)
            for w in words:
                if w['top'] > min_top + 60:
                    continue
                t = w['text']
                if t.startswith('200m'):    col_200  = w['x0']; header_y_max = max(header_y_max or 0, w['top'] + 10)
                elif t.startswith('1000m'): col_1000 = w['x0']; header_y_max = max(header_y_max or 0, w['top'] + 10)
                elif t.startswith('Pointsrac'): col_punt = w['x0']; header_y_max = max(header_y_max or 0, w['top'] + 10)
            # Skater-list-header alleen geldig als alle 3 kolommen op deze
            # pagina-header zijn gevonden — daarmee onderscheid van een
            # tabel die "Cadet Ladies 1000m" als tekst-cel bevat.
            page_heeft_skater_header = (
                col_200 is not None and col_1000 is not None
                and col_punt is not None and header_y_max is not None
            )
            if page_heeft_skater_header:
                in_skater_list = True
                skater_list_pages.add(pidx)
            if not in_skater_list:
                continue
            # Clustering met tolerantie ~3px ipv round(): de snr+naam staan
            # vaak op top=33.4 en de cat op 33.8 — round() splitst dat in
            # 33/34 wat de rij in tweeën breekt en de cat-match faalt.
            data_words = [
                w for w in words
                if not (header_y_max and w['top'] < header_y_max)
            ]
            data_words.sort(key=lambda w: w['top'])
            lines_clustered: list[list] = []
            for w in data_words:
                if lines_clustered and abs(w['top'] - lines_clustered[-1][0]['top']) < 3:
                    lines_clustered[-1].append(w)
                else:
                    lines_clustered.append([w])
            for rij_unsorted in lines_clustered:
                rij = sorted(rij_unsorted, key=lambda w: w['x0'])
                # Data-regel: begint met startnr
                if not rij or not rij[0]['text'].isdigit():
                    continue
                snr = int(rij[0]['text'])
                # Cat = woord dat aan regex matcht
                cat_idx = None
                for i, w in enumerate(rij):
                    if CAT_PATT.match(w['text']):
                        cat_idx = i
                        break
                if cat_idx is None:
                    continue
                cat = rij[cat_idx]['text']
                naam = ' '.join(w['text'] for w in rij[1:cat_idx]).strip()
                # Ticks ✕/✗ → bepaal kolom op basis van x0
                afstanden: set[str] = set()
                for w in rij:
                    if w['text'] in ('✕', '✗', '×'):
                        x = w['x0']
                        if col_200 is None:
                            continue
                        dists = [
                            (abs(x - col_200),  '200m'),
                            (abs(x - col_1000), '1000m'),
                            (abs(x - col_punt), 'puntenkoers'),
                        ]
                        nearest = min(dists, key=lambda d: d[0])[1]
                        afstanden.add(nearest)
                key = normaliseer_naam(naam)
                if key in pdf_d:
                    continue  # eerste hit wint
                pdf_d[key] = {
                    'snr':  snr,
                    'naam': naam.strip(),
                    'cat':  cat,
                    'club': '',
                    'afstanden': afstanden,
                    'sectie': 'bevestigd',
                }
    # ── "Geen transponder geregistreerd"-sectie op p1 parsen ─────────
    # Aparte info-stroom van de Deelnemerslijst — markeert per snr+naam
    # dat in InlineComp geen transponder geregistreerd staat. Wordt
    # later vergeleken met Excel kolom P ('transponder').
    geen_transponder_pdf: list[dict] = []
    with pdfplumber.open(path) as doc:
        gevonden_header = False
        for page in doc.pages:
            t = page.extract_text() or ''
            for line in t.splitlines():
                low = line.lower()
                if 'geen transponder geregistreerd' in low:
                    gevonden_header = True
                    continue
                if not gevonden_header:
                    continue
                # Eindig bij de volgende sectie-header
                if any(s in low for s in (
                        'race-groepen', 'overzicht', 'deelnemerslijst',
                        '# naam cat transp')):
                    if 'geen transponder' in low:
                        continue
                    if line.strip().startswith('# Naam'):
                        continue  # tabel-header in deze sectie zelf
                    gevonden_header = False
                    continue
                m = re.match(r'^\s*(\d{1,4})\s+(.+?)\s+([DH](?:KA|JA|JB|SA|SB|SJ|P[1-9]))\s*$',
                             line.strip())
                if m:
                    geen_transponder_pdf.append({
                        'snr':  int(m.group(1)),
                        'naam': m.group(2).strip(),
                        'cat':  m.group(3),
                    })

    # Pre-secties (p1 + p2-top: Niet aanwezig, Door organisatie toegevoegd,
    # Niet bevestigd, Race-groepen overzicht) worden NIET geparset — Geert
    # wil alleen de echte Deelnemerslijst-tabel die ná de "Deelnemerslijst
    # nnn deelnemers"-header begint.
    return list(pdf_d.values()), geen_transponder_pdf


# ── Matching ─────────────────────────────────────────────────────────────
def match_lijsten(pdf_lijst: list[dict], xls_lijst: list[dict], fuzzy_drempel: int = 85):
    """
    Bouwt:
      exact_paren:  (pdf, xls) waar snr én naam exact matchen
      fuzzy_paren:  (pdf, xls, score) waar naam ≥ drempel maar geen exact
      snr_verschil: (pdf, xls) waar naam exact matcht maar snr verschilt
      alleen_pdf:   rijen niet matched
      alleen_xls:   rijen niet matched
    """
    used_xls = set()
    exact_paren, fuzzy_paren, snr_verschil = [], [], []
    alleen_pdf = []

    # Index excel-rijen per (snr, norm-naam) en per norm-naam
    xls_by_snr_naam = {}
    xls_by_naam = {}
    for i, x in enumerate(xls_lijst):
        n = normaliseer_naam(x['naam'])
        if x['snr'] is not None:
            xls_by_snr_naam[(x['snr'], n)] = i
        xls_by_naam.setdefault(n, []).append(i)

    for p in pdf_lijst:
        n = normaliseer_naam(p['naam'])
        # 1. exact (snr + naam)
        idx = xls_by_snr_naam.get((p['snr'], n))
        if idx is not None and idx not in used_xls:
            exact_paren.append((p, xls_lijst[idx]))
            used_xls.add(idx)
            continue
        # 2. exact naam met ander snr
        cand_naam = [i for i in xls_by_naam.get(n, []) if i not in used_xls]
        if cand_naam:
            i = cand_naam[0]
            x = xls_lijst[i]
            if x['snr'] is not None and p['snr'] is not None and x['snr'] != p['snr']:
                snr_verschil.append((p, x))
            else:
                exact_paren.append((p, x))
            used_xls.add(i)
            continue
        # 3. fuzzy naam-match
        beste, beste_score = None, 0
        for i, x in enumerate(xls_lijst):
            if i in used_xls: continue
            s = fuzzy_score(p['naam'], x['naam'])
            if s > beste_score:
                beste_score = s
                beste = i
        if beste is not None and beste_score >= fuzzy_drempel:
            fuzzy_paren.append((p, xls_lijst[beste], beste_score))
            used_xls.add(beste)
            continue
        alleen_pdf.append(p)

    alleen_xls = [x for i, x in enumerate(xls_lijst) if i not in used_xls]
    return exact_paren, fuzzy_paren, snr_verschil, alleen_pdf, alleen_xls


# ── PDF-rapport ──────────────────────────────────────────────────────────
BLAUW = colors.HexColor('#1a3a5c')
GRIJS = colors.HexColor('#666666')


def maak_rapport(pdf_pad: Path, xls_pad: Path,
                 pdf_lijst, xls_lijst, exact, fuzzy, snr_verschil,
                 alleen_pdf, alleen_xls, output: Path,
                 geen_transp_pdf: list[dict] | None = None):
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle('h1', parent=styles['Heading1'],
                        textColor=BLAUW, fontSize=14, spaceAfter=2)
    sub = ParagraphStyle('sub', parent=styles['Normal'],
                         fontSize=9, textColor=GRIJS, spaceAfter=8)
    h2 = ParagraphStyle('h2', parent=styles['Heading2'],
                        textColor=BLAUW, fontSize=11,
                        spaceBefore=10, spaceAfter=4,
                        keepWithNext=1)
    normal = styles['Normal']
    # Variant voor sub-koppen die ook NIET op een eigen pagina mogen
    # eindigen (bv. "7a. Alleen in InlineComp ..." gevolgd door tabel).
    sub_kop = ParagraphStyle('sub_kop', parent=normal, keepWithNext=1)
    note = ParagraphStyle('note', parent=normal, fontSize=8, textColor=GRIJS)
    nt_kop = ParagraphStyle('nt_kop', parent=normal, fontSize=9, leading=11)

    doc = SimpleDocTemplate(str(output), pagesize=A4,
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=15*mm, bottomMargin=15*mm,
                            title='Deelnemerslijst-diff')
    story = []
    story.append(Paragraph('Diff-rapport deelnemers', h1))
    pdf_stand = _extract_pdf_datum(pdf_pad)
    pdf_label = (f'InlineComp dd {pdf_stand}'
                 if pdf_stand else f'InlineComp <i>{pdf_pad.name}</i>')
    xls_label = f"Excel '{xls_pad.name}'"
    story.append(Paragraph(
        f'{pdf_label} vs {xls_label} · '
        f'Gegenereerd op {datetime.now():%d-%m-%Y %H:%M}', sub))

    # ── Samenvattings-tabel ──
    n_pdf = len(pdf_lijst)
    n_xls = len(xls_lijst)
    excel_zonder_snr = sum(1 for x in xls_lijst if x['snr'] is None)
    afstand_mismatches = sum(
        1 for paar in (exact + [(p, x) for p, x, _ in fuzzy])
        if paar[0]['afstanden'] != paar[1]['afstanden']
    )

    samenvatting = [
        ['Categorie', 'InlineComp', 'Organisatie'],
        ['Totaal aantal deelnemers',          f'{n_pdf} (PDF)', f'{n_xls} (XLS)'],
        ['Exact match (naam + snr + cat)',    str(len(exact)), ''],
        ['Naam-spelling-verschillen (fuzzy)', str(len(fuzzy)), ''],
        ['Alleen in PDF (uit InlineComp)',    str(len(alleen_pdf)), ''],
        ['Alleen in XLS (uit Excel)',         '', str(len(alleen_xls))],
        ['Afstand-mismatches',                str(afstand_mismatches), ''],
        ['Startnummer-verschillen',           str(len(snr_verschil)), ''],
        ['Excel zonder ingevuld startnr',     '', str(excel_zonder_snr)],
    ]
    t = Table(samenvatting, colWidths=[85*mm, 45*mm, 45*mm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), BLAUW),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 9),
        ('GRID',       (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ('ALIGN',      (1, 1), (-1, -1), 'CENTER'),
        ('LEFTPADDING',(0,0),(-1,-1),5), ('RIGHTPADDING',(0,0),(-1,-1),5),
    ]))
    story.append(t)
    story.append(Spacer(1, 4*mm))

    # ── Sectie 1: Alleen in Excel ──
    story.append(Paragraph(
        f"1. Deelnemer/ster alleen in {xls_label}, "
        f"niet in {pdf_label}", h2))
    if not alleen_xls:
        story.append(Paragraph('(geen)', normal))
    else:
        data = [['Snr', 'Naam', 'Cat', 'Club']]
        for x in alleen_xls:
            data.append([
                f'#{x["snr"]}' if x['snr'] else '—',
                x['naam'], x['cat'], (x['club'] or '')[:30],
            ])
        story.append(_basis_tabel(data, [16*mm, 60*mm, 18*mm, 82*mm]))

    # ── Sectie 2: Alleen in InlineComp ──
    story.append(Paragraph(
        f"2. Deelnemer/ster alleen in {pdf_label}, "
        f"niet vermeld in {xls_label}", h2))
    if not alleen_pdf:
        story.append(Paragraph('(geen)', normal))
    else:
        data = [['Snr', 'Naam', 'Cat', 'Afstanden']]
        for p in alleen_pdf:
            data.append([
                f'#{p["snr"]}',
                p['naam'], p['cat'],
                ', '.join(sorted(p['afstanden'])) or '—',
            ])
        story.append(_basis_tabel(data, [16*mm, 60*mm, 18*mm, 82*mm]))

    # ── Sectie 3: Naam-verschillen (fuzzy) ──
    story.append(Paragraph(
        '3. Naam-spelling verschilt tussen InlineComp en Excel '
        '(zelfde rijder, andere notatie)', h2))
    if not fuzzy:
        story.append(Paragraph('(geen)', normal))
    else:
        data = [['Snr', 'InlineComp', 'Excel', 'Match']]
        for p, x, score in sorted(fuzzy, key=lambda t: (t[0]['snr'] or 9999)):
            data.append([
                f'#{p["snr"]}',
                p['naam'], x['naam'], f'{score}%'
            ])
        story.append(_basis_tabel(data, [16*mm, 60*mm, 60*mm, 18*mm]))

    # ── Sectie 4: Afstand-verschillen (tabel-vorm) ──
    story.append(Paragraph(
        '4. Afstand-inschrijving verschilt tussen InlineComp en Excel', h2))
    matched_paren = exact + [(p, x) for p, x, _ in fuzzy]
    afst_diffs = [(p, x) for p, x in matched_paren
                  if p['afstanden'] != x['afstanden']]
    if not afst_diffs:
        story.append(Paragraph('(geen)', normal))
    else:
        # Eén rij per (rijder, afstand-met-verschil). De laatste twee
        # kolommen tonen of de afstand in elke bron is aangevinkt.
        # Korte kolom-headers — de context "InlineComp vs Excel" staat al in
        # de sectie-titel boven de tabel. pdf_label kan <i>...</i>-HTML
        # bevatten voor Paragraph-gebruik en is te lang voor smalle kolommen.
        data = [['Snr', 'Naam', 'Cat', 'Afstand',
                 'In InlineComp', 'In Excel']]
        for p, x in sorted(afst_diffs, key=lambda t: (t[0]['snr'] or 9999)):
            alle = sorted(p['afstanden'] | x['afstanden'])
            for a in alle:
                in_pdf = a in p['afstanden']
                in_xls = a in x['afstanden']
                if in_pdf == in_xls:
                    continue
                data.append([
                    f'#{p["snr"]}',
                    p['naam'],
                    p['cat'],
                    a,
                    'ja' if in_pdf else 'nee',
                    'ja' if in_xls else 'nee',
                ])
        # Totaal 176 mm = ~content-breedte (A4 minus 15 mm margins).
        story.append(_basis_tabel(
            data, [12*mm, 45*mm, 14*mm, 25*mm, 40*mm, 40*mm]))

    # ── Sectie 5: Startnummer-verschillen ──
    story.append(Paragraph(
        '5. Startnummer verschilt tussen InlineComp en Excel', h2))
    snr_in_excel_zonder_match = [
        x for x in alleen_xls if x['snr'] is None
    ]
    if not snr_verschil and not snr_in_excel_zonder_match:
        story.append(Paragraph('(geen)', normal))
    else:
        data = [['Naam', 'InlineComp', 'Excel']]
        for p, x in sorted(snr_verschil, key=lambda t: (t[0]['snr'] or 9999)):
            data.append([p['naam'], f'#{p["snr"]}', f'#{x["snr"]}'])
        # Excel-rijen waar snr ontbreekt maar de naam wel matched elders
        # (deze worden niet automatisch als snr-verschil getoond — laten we
        # de echte rapportage simpel houden)
        story.append(_basis_tabel(data, [80*mm, 45*mm, 45*mm]))

    # ── Sectie 6: Missende geslacht-info in Excel ──
    missende_gesl = [x for x in xls_lijst if x.get('geslacht_ontbreekt')]
    story.append(Paragraph(
        '6. Geslacht-info ontbreekt in Excel', h2))
    if not missende_gesl:
        story.append(Paragraph('(geen)', normal))
    else:
        data = [['Snr', 'Naam', 'Cat']]
        for x in sorted(missende_gesl, key=lambda x: x['snr'] or 9999):
            data.append([
                f'#{x["snr"]}' if x['snr'] else '—',
                x['naam'], x['cat'],
            ])
        story.append(_basis_tabel(data, [16*mm, 80*mm, 80*mm]))

    # ── Sectie 7: Transponder-verschillen ──
    if geen_transp_pdf:
        # Bouw naam-keys voor vergelijking. Match met fuzzy_score (≥85)
        # ipv exact, anders raken "Francisca" vs "Fracisca" of "Bech-Ryding"
        # vs "Bech- Ryding" dubbel in 7a én 7b.
        pdf_geen = {normaliseer_naam(x['naam']): x for x in geen_transp_pdf}
        xls_geen = {normaliseer_naam(x['naam']): x for x in xls_lijst
                    if x.get('geen_transponder')}
        in_beide_geen: list[str] = []
        alleen_pdf_geen: list[str] = []
        gematched_xls: set[str] = set()
        for k_pdf in sorted(pdf_geen):
            naam_pdf = pdf_geen[k_pdf]['naam']
            beste_k, beste_s = None, 0
            for k_xls in xls_geen:
                if k_xls in gematched_xls:
                    continue
                s = fuzzy_score(naam_pdf, xls_geen[k_xls]['naam'])
                if s > beste_s:
                    beste_s, beste_k = s, k_xls
            if beste_s >= 85 and beste_k:
                in_beide_geen.append(k_pdf)
                gematched_xls.add(beste_k)
            else:
                alleen_pdf_geen.append(k_pdf)
        alleen_xls_geen = sorted(k for k in xls_geen if k not in gematched_xls)

        story.append(Paragraph(
            f'7. Transponder-registratie verschilt tussen InlineComp en Excel '
            f'({len(in_beide_geen)} match · {len(alleen_pdf_geen)} alleen '
            f'InlineComp · {len(alleen_xls_geen)} alleen Excel)', h2))
        story.append(Paragraph(
            f'Bron: sectie "Geen transponder geregistreerd" in {pdf_label} '
            f'vs Excel-kolom P "transponder" in {xls_label}.', note))

        if alleen_pdf_geen:
            story.append(Paragraph(
                f'<b>7a.</b> Zonder transponder volgens {pdf_label}, '
                f'wél transponder volgens {xls_label} '
                f'({len(alleen_pdf_geen)})',
                sub_kop))
            data = [['Snr', 'Naam', 'Cat']]
            for k in alleen_pdf_geen:
                p = pdf_geen[k]
                data.append([f'#{p["snr"]}', p['naam'], p['cat']])
            story.append(_basis_tabel(data, [16*mm, 80*mm, 80*mm]))
            story.append(Spacer(1, 2*mm))

        if alleen_xls_geen:
            story.append(Paragraph(
                f'<b>7b.</b> Zonder transponder volgens {xls_label}, '
                f'wél transponder volgens {pdf_label} '
                f'({len(alleen_xls_geen)})',
                sub_kop))
            data = [['Snr', 'Naam', 'Cat']]
            for k in alleen_xls_geen:
                x = xls_geen[k]
                data.append([
                    f'#{x["snr"]}' if x['snr'] else '—',
                    x['naam'], x['cat'],
                ])
            story.append(_basis_tabel(data, [16*mm, 80*mm, 80*mm]))

    # ── Actie-checklist ──
    story.append(Spacer(1, 4*mm))
    story.append(Paragraph('Actie-checklist voor de organisatie', h2))
    acties = []
    if alleen_pdf:
        acties.append(
            f'Schrap of voeg in de organisatie-Excel de {len(alleen_pdf)} '
            f'alleen-in-InlineComp rijder(s) toe')
    if alleen_xls:
        acties.append(
            f'Voeg in InlineComp de {len(alleen_xls)} alleen-in-Excel '
            f'rijder(s) handmatig toe of importeer ze')
    if fuzzy:
        acties.append(
            'Corrigeer spelling-typos in organisatie-Excel zodat beide lijsten matchen op naam')
    if afst_diffs:
        acties.append(
            f'Verifieer en bevestig de {len(afst_diffs)} afstand-mismatches')
    if snr_verschil:
        acties.append(
            f'Verifieer {len(snr_verschil)} startnummer-verschil(len)')
    if excel_zonder_snr:
        acties.append(
            f'Vul de {excel_zonder_snr} ontbrekende startnummers in de Excel aan')
    if missende_gesl:
        acties.append(
            f'Vul het geslacht aan voor {len(missende_gesl)} rijders in de Excel')
    if not acties:
        acties.append('Alles in orde — geen acties vereist.')
    for a in acties:
        story.append(Paragraph(f'☐ {a}', normal))

    # ── Bronbestanden ──
    story.append(Spacer(1, 4*mm))
    story.append(Paragraph('Bronbestanden:', note))
    story.append(Paragraph(
        f'· InlineComp: {pdf_pad.name} ({n_pdf} skaters)', note))
    story.append(Paragraph(
        f'· Organisatie: {xls_pad.name} ({n_xls} aanmeldingen)', note))

    doc.build(story)


def _basis_tabel(data, col_widths):
    t = Table(data, repeatRows=1, colWidths=col_widths)
    t.hAlign = 'LEFT'  # alle tabellen vanaf de linker-marge, niet centered
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), BLAUW),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 8),
        ('GRID',       (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ('VALIGN',     (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING',(0,0),(-1,-1),4), ('RIGHTPADDING',(0,0),(-1,-1),4),
    ]))
    return t


def main():
    args = sys.argv[1:]
    # --sheet=<naam> filteren uit args
    sheet_arg = None
    overig = []
    for a in args:
        if a.startswith('--sheet='):
            sheet_arg = a.split('=', 1)[1]
        else:
            overig.append(a)
    if len(overig) < 2:
        print(__doc__); sys.exit(1)
    pdf_path = Path(overig[0])
    xls_path = Path(overig[1])
    out_path = Path(overig[2]) if len(overig) > 2 else \
               Path.home() / 'Downloads' / 'OH850_diff_rapport_v24.pdf'
    pdf_lijst, geen_transp_pdf = parse_pdf(pdf_path)
    xls_lijst, _ = parse_excel(xls_path, sheet=sheet_arg)
    exact, fuzzy, snr_verschil, alleen_pdf, alleen_xls = \
        match_lijsten(pdf_lijst, xls_lijst)
    maak_rapport(pdf_path, xls_path, pdf_lijst, xls_lijst,
                 exact, fuzzy, snr_verschil, alleen_pdf, alleen_xls,
                 out_path, geen_transp_pdf=geen_transp_pdf)
    print(f'Output: {out_path}')
    print(f'  PDF: {len(pdf_lijst)} deelnemers')
    print(f'  XLS: {len(xls_lijst)} aanmeldingen')
    print(f'  exact match:       {len(exact)}')
    print(f'  fuzzy (naam):      {len(fuzzy)}')
    print(f'  snr-verschil:      {len(snr_verschil)}')
    print(f'  alleen PDF:        {len(alleen_pdf)}')
    print(f'  alleen XLS:        {len(alleen_xls)}')


if __name__ == '__main__':
    main()
